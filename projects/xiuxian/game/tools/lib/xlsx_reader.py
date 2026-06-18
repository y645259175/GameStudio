"""xlsx 读取（openpyxl read_only，从不写）。

约定的表头格式：
  第 1 行：策划友好的中文表头（含"备注"列）
  第 2 行：英文字段名（程序读这行）
  第 3 行起：数据

注：A1 也算第 1 行。所谓"第 1 行 / 第 2 行"按 1-based。
"""
from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore

from .errors import ConvertError


REMARK_HEADER_CN = "备注"


@dataclass
class SheetData:
    """读取后的扁平数据（已剔除备注列、已应用承上规则）。"""
    sheet: str
    field_names: list[str]              # 英文字段名顺序（剔除备注列后）
    cn_headers: list[str]               # 中文表头顺序（同上对齐）
    rows: list[dict[str, Any]] = field(default_factory=list)
    row_xlsx_lines: list[int] = field(default_factory=list)  # 每行对应 xlsx 真实行号（用于错误定位）


def read_sheet(
    xlsx_path: Path,
    sheet_name: str,
    inherit_when_blank: list[str] | None = None,
) -> SheetData:
    """读取一个 sheet。

    返回 SheetData，备注列已剔除，承上规则已应用。
    """
    inherit_when_blank = inherit_when_blank or []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ConvertError(
            f"sheet not found: '{sheet_name}'. available: {wb.sheetnames}",
            xlsx=str(xlsx_path),
        )
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=False)

    # 第 1 行：中文表头
    try:
        cn_row = next(rows_iter)
    except StopIteration:
        raise ConvertError("empty sheet", xlsx=str(xlsx_path), sheet=sheet_name)
    cn_headers_raw = [_cell_str(c) for c in cn_row]

    # 第 2 行：英文字段名
    try:
        en_row = next(rows_iter)
    except StopIteration:
        raise ConvertError(
            "missing english field-name row (row 2)",
            xlsx=str(xlsx_path), sheet=sheet_name,
        )
    en_field_raw = [_cell_str(c) for c in en_row]

    # 计算保留列：cn_header != "备注" 且 en_field 非空
    keep_cols: list[int] = []
    field_names: list[str] = []
    cn_headers: list[str] = []
    for idx, (cn, en) in enumerate(zip(cn_headers_raw, en_field_raw)):
        if cn.strip() == REMARK_HEADER_CN:
            continue
        if not en.strip():
            continue
        keep_cols.append(idx)
        field_names.append(en.strip())
        cn_headers.append(cn.strip())

    # 主键检测重复字段名
    if len(set(field_names)) != len(field_names):
        raise ConvertError(
            f"duplicate field names: {field_names}",
            xlsx=str(xlsx_path), sheet=sheet_name, row=2,
        )

    sd = SheetData(sheet=sheet_name, field_names=field_names, cn_headers=cn_headers)

    # 数据行 row 3+
    last_inherit_values: dict[str, Any] = {}
    xlsx_row = 2
    for row_cells in rows_iter:
        xlsx_row += 1
        # 全行空跳过
        if all(_cell_str(row_cells[i]) == "" for i in keep_cols if i < len(row_cells)):
            continue
        rec: dict[str, Any] = {}
        for col_idx, fname in zip(keep_cols, field_names):
            v_raw = row_cells[col_idx].value if col_idx < len(row_cells) else None
            v = _normalize(v_raw)
            if v == "" and fname in inherit_when_blank:
                v = last_inherit_values.get(fname, "")
            rec[fname] = v
            if fname in inherit_when_blank and v != "":
                last_inherit_values[fname] = v
        sd.rows.append(rec)
        sd.row_xlsx_lines.append(xlsx_row)
    wb.close()
    return sd


def _cell_str(cell) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value)


def _normalize(v: Any) -> Any:
    """空白字符串 → ''；保留 int/float/bool；其它 str 化"""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float, bool)):
        return v
    return str(v).strip()
