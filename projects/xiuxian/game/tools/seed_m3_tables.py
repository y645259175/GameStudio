"""一次性脚本：生成 M3 第一画面所需的 3 张配表 xlsx。
数据来源：GDD-04 / GDD-05 / GDD-06。

运行：python game/tools/seed_m3_tables.py
位置：xlsx 输出到 game/data/table/<名>.xlsx
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook  # type: ignore

THIS = Path(__file__).resolve()
TABLE_DIR = THIS.parent.parent / "data" / "table"
TABLE_DIR.mkdir(parents=True, exist_ok=True)


def write_sheet(wb: Workbook, sheet_name: str, headers_cn: list[str], headers_en: list[str], rows: list[list]):
    ws = wb.create_sheet(sheet_name) if sheet_name not in wb.sheetnames else wb[sheet_name]
    ws.append(headers_cn)
    ws.append(headers_en)
    for r in rows:
        ws.append(r)


# ---------------------------------------------------------------------------
# 建筑配置.xlsx
# ---------------------------------------------------------------------------
def gen_building():
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    headers_cn = ["行ID", "建筑ID", "等级", "中文名", "图标路径", "建造月数", "灵石成本", "modifier类型", "modifier数值", "弟子容量", "中文描述", "备注"]
    headers_en = ["row_id", "building_id", "level", "name_cn", "icon_path", "build_months", "cost_spirit_stone", "modifier_kind", "modifier_value", "capacity", "desc_cn", ""]

    # 5 建筑 × 3 级 = 15 行（GDD-05 §3.2 / §4.5 数值）
    rows = [
        # 修炼塔（GDD-05 §3.2.3 修炼速度 / §7.5 容量 2/3/5）
        ["cultivation_tower_lv1", "cultivation_tower", 1, "修炼塔", "res://art/m3/buildings/cultivation_tower.png", 1, 200, "qi_acceleration", 20, 2, "提升驻塔弟子修炼速度，lv1 +20%", ""],
        ["cultivation_tower_lv2", "",                  2, "",       "",                                              2, 500, "",                  35, 3, "lv2 +35%",                              ""],
        ["cultivation_tower_lv3", "",                  3, "",       "",                                              2, 1000, "",                 50, 5, "lv3 +50%",                              ""],
        # 藏经阁（GDD-05 §3.2.4 悟性 / §7.5 容量 3/6/10）
        ["library_lv1",           "library",           1, "藏经阁", "res://art/m3/buildings/library.png",            1, 250, "insight_per_month",  1, 3, "驻阁修炼者每月 +1 悟性",                ""],
        ["library_lv2",           "",                  2, "",       "",                                              2, 600, "",                   2, 6, "+2 悟性/月",                            ""],
        ["library_lv3",           "",                  3, "",       "",                                              2, 1200, "",                  3, 10, "+3 悟性/月",                           ""],
        # 丹房（GDD-05 §3.2 / GDD-04 §6 炼丹）
        ["alchemy_room_lv1",      "alchemy_room",      1, "丹房",   "res://art/m3/buildings/alchemy_room.png",       1, 300, "forge_speed",        20, 1, "炼丹任务速度 +20%",                      ""],
        ["alchemy_room_lv2",      "",                  2, "",       "",                                              2, 700, "",                  35, 2, "+35%",                                   ""],
        ["alchemy_room_lv3",      "",                  3, "",       "",                                              2, 1500, "",                 50, 3, "+50%",                                   ""],
        # 弟子居所（GDD-05 §3.2.2 housing_capacity 容量上限，预建）
        ["disciple_dorm_lv1",     "disciple_dorm",     1, "弟子居所", "res://art/m3/buildings/dormitory.png",         1, 100, "housing_capacity",   6, 0, "宗门可容纳 6 名弟子",                    ""],
        ["disciple_dorm_lv2",     "",                  2, "",         "",                                            1, 250, "",                  10, 0, "可容纳 10 名",                          ""],
        ["disciple_dorm_lv3",     "",                  3, "",         "",                                            2, 600, "",                  15, 0, "可容纳 15 名",                          ""],
        # 主殿（GDD-05 §3.2.1 声望基准 + 槽位解锁器，预建）
        ["main_hall_lv1",         "main_hall",         1, "主殿",     "res://art/m3/buildings/main_hall.png",        1, 0,   "reputation_gain",    1.0, 0, "宗门核心，招收入口与威望象征",          ""],
        ["main_hall_lv2",         "",                  2, "",         "",                                            2, 800, "",                   2.0, 0, "提升招收质量上限",                      ""],
        ["main_hall_lv3",         "",                  3, "",         "",                                            2, 2000, "",                  3.0, 0, "宗门威望大涨",                          ""],
    ]
    write_sheet(wb, "建筑等级", headers_cn, headers_en, rows)

    out = TABLE_DIR / "建筑配置.xlsx"
    wb.save(out)
    print(f"[ok] {out.name}: {len(rows)} rows")


# ---------------------------------------------------------------------------
# 资源配置.xlsx
# ---------------------------------------------------------------------------
def gen_resource():
    wb = Workbook()
    wb.remove(wb.active)

    headers_cn = ["资源ID", "中文名", "图标路径", "类别", "上限", "描述", "备注"]
    headers_en = ["resource_id", "name_cn", "icon_path", "category", "cap", "desc_cn", ""]
    rows = [
        ["spirit_stone",   "灵石",     "res://art/m3/icons/spirit_stone.png", "currency",   0,    "宗门基础货币，建造/招募/炼丹皆需要", ""],
        ["spirit_herb",    "灵草",     "res://art/m3/icons/herb.png",         "material",   9999, "炼丹原料，野外历练采集",            ""],
        ["beast_blood",    "兽血",     "",                                    "material",   999,  "炼丹原料，历练战利品",              ""],
        ["demon_core",     "妖丹",     "",                                    "material",   999,  "妖兽内丹，高阶炼丹原料",            ""],
        ["fire_essence",   "火精",     "",                                    "material",   999,  "火属性精华，灵根丹原料",            ""],
        ["millennium_herb","千年灵芝", "",                                    "material",   99,   "稀有药材，寿元丹原料",              ""],
        ["mortal_pill",    "凡品丹药", "res://art/m3/icons/pill_mortal.png",  "consumable", 999,  "回复体力 / 消除轻伤 / 短期 buff",   ""],
        ["yellow_pill",    "黄品丹药", "res://art/m3/icons/pill_yellow.png",  "consumable", 999,  "中阶丹药",                            ""],
        ["mystic_pill",    "玄品丹药", "res://art/m3/icons/pill_mystic.png",  "consumable", 99,   "高阶丹药",                            ""],
        ["breakthrough_pill_qi", "炼气突破丹", "res://art/m3/icons/pill_breakthrough.png", "consumable", 99, "炼气境突破检定时消耗，提升基础概率", ""],
        ["sect_token",     "宗门令牌", "",                                    "treasure",   1,    "宗门权柄象征",                        ""],
    ]
    write_sheet(wb, "资源", headers_cn, headers_en, rows)

    out = TABLE_DIR / "资源配置.xlsx"
    wb.save(out)
    print(f"[ok] {out.name}: {len(rows)} rows")


# ---------------------------------------------------------------------------
# 境界曲线.xlsx
# ---------------------------------------------------------------------------
def gen_realm_curve():
    wb = Workbook()
    wb.remove(wb.active)

    headers_cn = ["主键", "大境", "大境中文", "小境", "升级所需经验", "基础修炼速度", "突破基础分", "突破难度系数", "突破后寿元+月", "战力倍率", "备注"]
    headers_en = ["realm_sub", "realm_id", "realm_cn", "sub_level", "exp_threshold", "base_speed", "breakthrough_base_score", "breakthrough_difficulty", "lifespan_bonus_months", "combat_power_multiplier", ""]

    # GDD-06 §6.2 / §7.1 数值
    # 大境跳变倍率（combat_power）：凡=1 / 炼气=2 / 筑基=4 / 金丹=8 / 元婴=16 / 化神=32（GDD-02 §3.5 已删表的语义保留）
    # base_speed 每大境跳 ~1.5x（GDD-06 §6.2 估算）
    # exp_threshold 每小境 100 起步，跨大境 ×3（数值占位，可调）
    REALMS = [
        ("mortal",      "凡人", 1.0,   1),
        ("qi",          "炼气", 1.0,   2),
        ("foundation",  "筑基", 1.5,   4),
        ("golden",      "金丹", 2.5,   8),
        ("yuanying",    "元婴", 4.0,  16),
        ("huashen",     "化神", 6.5,  32),
    ]
    BASE_SPEED_MORTAL = 5.0   # 凡人每月修为
    EXP_BASE = {  # 该大境基础经验阈值
        "mortal": 50.0, "qi": 100.0, "foundation": 300.0,
        "golden": 900.0, "yuanying": 2700.0, "huashen": 8100.0,
    }
    BREAK_BASE = {"mortal": 0, "qi": 20, "foundation": 35, "golden": 60, "yuanying": 100, "huashen": 200}
    BREAK_DIFF = {"mortal": 50, "qi": 100, "foundation": 200, "golden": 400, "yuanying": 800, "huashen": 1600}
    LIFESPAN_BONUS = {"mortal": 0, "qi": 1080, "foundation": 1080, "golden": 1080, "yuanying": 1080, "huashen": 1080}  # 90 年/120 月 = 不一致；按 80-100 年取 1080 月（90 年）

    rows = []
    for realm_id, realm_cn, speed_mult, power_mult in REALMS:
        max_sub = 1 if realm_id == "mortal" else 9
        for sub in range(1, max_sub + 1):
            row_id = f"{realm_id}_{sub}"
            exp = EXP_BASE[realm_id] * (1.0 + 0.15 * (sub - 1))   # 同大境内每小境 +15% 阈值
            speed = BASE_SPEED_MORTAL * speed_mult * (1.0 + 0.05 * (sub - 1))
            # 大境最后一层(9)给突破参数，其余 0
            is_top = (sub == 9)
            rows.append([
                row_id,
                realm_id if sub == 1 else "",     # 承上
                realm_cn if sub == 1 else "",
                sub,
                round(exp, 2),
                round(speed, 2),
                BREAK_BASE[realm_id] if is_top else 0,
                BREAK_DIFF[realm_id] if is_top else 0,
                LIFESPAN_BONUS[realm_id] if is_top else 0,
                round(float(power_mult), 2),
                ""
            ])

    write_sheet(wb, "境界曲线", headers_cn, headers_en, rows)

    out = TABLE_DIR / "境界曲线.xlsx"
    wb.save(out)
    print(f"[ok] {out.name}: {len(rows)} rows")


# ---------------------------------------------------------------------------
# 炼丹配方.xlsx（GDD-05 §5.3）
# ---------------------------------------------------------------------------
def gen_alchemy():
    wb = Workbook()
    wb.remove(wb.active)
    headers_cn = ["配方ID", "中文名", "产物图标", "材料1ID", "材料1数量", "材料2ID", "材料2数量",
                  "所需炼丹值", "所需丹房等级", "炼制月数", "成功率", "产物ID", "产物数量", "描述", "备注"]
    headers_en = ["recipe_id", "name_cn", "icon_path", "material1_id", "material1_count", "material2_id", "material2_count",
                  "required_alchemy_skill", "required_room_level", "duration_months", "success_rate", "output_id", "output_count", "desc_cn", ""]
    rows = [
        ["recipe_qi_pill",          "聚气丹",   "res://art/m3/icons/pill_mortal.png", "spirit_herb", 2, "",            0, 10, 1, 1, 0.90, "mortal_pill",          3, "闭关加速丹药，新手即可炼", ""],
        ["recipe_healing_pill",     "疗伤丹",   "res://art/m3/icons/pill_mortal.png", "spirit_herb", 1, "beast_blood", 1, 20, 1, 1, 0.80, "yellow_pill",          2, "治疗重伤", ""],
        ["recipe_breakthrough_pill","突破丹",   "res://art/m3/icons/pill_breakthrough.png", "spirit_herb", 3, "demon_core", 1, 40, 2, 2, 0.60, "breakthrough_pill_qi", 1, "突破成功率 +10%", ""],
        ["recipe_root_boost_fire",  "火灵根丹", "res://art/m3/icons/pill_mystic.png", "fire_essence", 2, "demon_core",  2, 60, 3, 2, 0.45, "mystic_pill",          1, "火灵根 +1", ""],
        ["recipe_lifespan_pill",    "寿元丹",   "res://art/m3/icons/pill_yellow.png", "millennium_herb", 1, "demon_core", 3, 80, 3, 3, 0.30, "yellow_pill",          1, "寿元 +5 年（M4 启用）", ""],
    ]
    write_sheet(wb, "炼丹配方", headers_cn, headers_en, rows)
    out = TABLE_DIR / "炼丹配方.xlsx"
    wb.save(out)
    print(f"[ok] {out.name}: {len(rows)} rows")


if __name__ == "__main__":
    gen_building()
    gen_resource()
    gen_realm_curve()
    gen_alchemy()
    print("[done] M3 配表 xlsx 生成完毕")
