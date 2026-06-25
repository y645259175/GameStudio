"""一次性脚本：生成 M3 历练事件配表。
数据来源：GDD-03 §2 事件系统最小可玩内容。

运行：python game/tools/seed_events.py
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook

THIS = Path(__file__).resolve()
TABLE_DIR = THIS.parent.parent / "data" / "table"
TABLE_DIR.mkdir(parents=True, exist_ok=True)


def ws_append(ws, headers_cn, headers_en, rows):
    ws.append(headers_cn)
    ws.append(headers_en)
    for r in rows:
        ws.append(r)


def gen():
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 事件模板（5 个历练事件 + 1 个系统事件）
    _hs = ["事件ID", "标题", "描述文本", "类别", "最小节点深度", "权重", "备注"]
    _en = ["event_id", "title_cn", "desc_cn", "category", "min_node_depth", "weight", ""]
    events = [
        # 第1节点（浅层）：轻量事件
        ["exp_forest_stroll", "林间漫步", "前方的树林灵气弥漫，隐约可见几株灵草在阳光下闪着微光。", "treasure", 0, 40, "第1节点高权"],
        ["exp_ancient_stele", "古碑残文", "一块半埋的石碑上刻着古老的符文，似乎记载着某种心法。", "story", 0, 30, ""],
        # 第2节点（中层）：选项事件
        ["exp_cave_discover", "洞窟抉择", "前方山壁有一个洞口，洞内隐隐有灵气发出，但也似乎有危险的气息。", "trial", 1, 50, "show_options 分支"],
        ["exp_spirit_beast", "灵兽袭扰", "一头受伤的赤鬃灵虎拦在路中央，对着你发出威严的低吼。", "battle", 1, 30, ""],
        # 第3节点（深层）：boss
        ["exp_boss_guardian", "山谷守卫", "一只巨大的碧鳞蟾蜍守在前往秘境的必经之路上，浑身散发着浓烈的妖气。", "battle", 2, 60, "boss 战"],
        # === 扩充事件（M3-3 ≥10 种）===
        # 浅层
        ["exp_herb_field", "灵草秘园", "一片隐蔽的山谷里长满了各色灵草，灵气扑面而来。", "treasure", 0, 35, ""],
        ["exp_old_hermit", "山中隐者", "一位白发老者在溪边垂钓，见你来到，微微一笑。", "encounter", 0, 25, ""],
        ["exp_spirit_spring", "灵泉涌动", "一汪清泉从石缝中涌出，泉水蕴含精纯灵气。", "treasure", 0, 30, ""],
        # 中层
        ["exp_ruins_choice", "古修遗府", "一座坍塌的洞府半掩在藤蔓之下，似有传承，亦有禁制。", "trial", 1, 45, "show_options"],
        ["exp_wandering_cultivator", "散修问道", "一名散修拦路，要与你切磋一番，赌注是身上的灵石。", "battle", 1, 35, ""],
        ["exp_poison_swamp", "瘴气沼泽", "前方是一片散发着紫色瘴气的沼泽，强行通过会损耗灵石。", "story", 1, 25, ""],
        # 深层
        ["exp_demon_ambush", "妖群围攻", "三只金瞳妖豹从林中扑出，将你团团围住！", "battle", 2, 40, "群战"],
        ["exp_treasure_vault", "秘境宝库", "秘境深处的宝库大门虚掩，里面金光闪烁。", "treasure", 2, 35, ""],
        # 系统事件
        ["__sys_extraction__", "撤离历练", "历练结束，收获归宗。", "system", 0, 0, "系统保留"],
    ]
    ws_append(wb.create_sheet("事件模板"), _hs, _en, events)

    # Sheet 2: 事件动作（最小可玩，只用 action_type+param）
    _ah = ["行ID", "事件ID", "顺序", "动作类型", "条件", "参数1", "参数2", "参数3", "备注"]
    _ae = ["row_id", "event_id", "order", "action_type", "condition", "param1", "param2", "param3", ""]
    actions = [
        # exp_forest_stroll：显示文本 → 给灵草
        ["exp_forest_stroll_1", "exp_forest_stroll", 1, "show_text", "", "前方的树林灵气弥漫，隐约可见几株灵草。", "", "", ""],
        ["exp_forest_stroll_2", "",                    2, "give_resource", "", "spirit_herb", "5", "", ""],
        # exp_ancient_stele：显示文本 → 给灵石+悟性flag
        ["exp_ancient_stele_1", "exp_ancient_stele", 1, "show_text", "", "古碑上刻着'大道无形，生育天地'，你静心参悟，若有所得。", "", "", ""],
        ["exp_ancient_stele_2", "",                   2, "give_resource", "", "spirit_stone", "300", "", ""],
        ["exp_ancient_stele_3", "",                   3, "set_flag", "", "got_insight", "1", "", ""],
        # exp_cave_discover：show_text → show_options（选探索=战斗+灵石 / 选绕路=灵草）
        ["exp_cave_discover_1", "exp_cave_discover", 1, "show_text", "", "洞口发出幽幽绿光，你听到了低沉的嘶鸣声。", "", "", ""],
        ["exp_cave_discover_2", "",                   2, "show_options", "", "opt_cave_explore", "opt_cave_bypass", "", ""],
        # 分支 A：探索 → 战斗 → 灵石
        ["exp_cave_discover_3", "",                   3, "show_text", "flag('choice') == 'opt_cave_explore'", "你握紧武器，迈入洞穴深处——一只石魔猿迎面扑来！", "", "", ""],
        ["exp_cave_discover_4", "",                   4, "start_battle", "flag('choice') == 'opt_cave_explore'", "", "", "", ""],
        ["exp_cave_discover_5", "",                   5, "show_text", "flag('battle_result') == 'win'", "石魔猿轰然倒地，洞内灵气四溢，你发现了不少灵石。", "", "", ""],
        ["exp_cave_discover_6", "",                   6, "give_resource", "flag('battle_result') == 'win'", "spirit_stone", "500", "", ""],
        ["exp_cave_discover_7", "",                   7, "show_text", "flag('battle_result') == 'lose'", "你且战且退，撤出洞穴，身上的灵石掉落了几颗...", "", "", ""],
        ["exp_cave_discover_8", "",                   8, "give_resource", "flag('battle_result') == 'lose'", "spirit_stone", "-200", "", ""],
        # 分支 B：绕路 → 灵草
        ["exp_cave_discover_9", "",                   9, "show_text", "flag('choice') == 'opt_cave_bypass'", "你选择安全为上，绕开洞穴，从山路采集了沿途的灵草。", "", "", ""],
        ["exp_cave_discover_10","",                   10, "give_resource", "flag('choice') == 'opt_cave_bypass'", "spirit_herb", "10", "", ""],
        # exp_spirit_beast：show_text → 战斗 → 灵草
        ["exp_spirit_beast_1",   "exp_spirit_beast",  1, "show_text", "", "赤鬃灵虎虽然是受伤之躯，但眼中依然透着猛兽的威严。", "", "", ""],
        ["exp_spirit_beast_2",   "",                   2, "start_battle", "", "", "", "", ""],
        ["exp_spirit_beast_3",   "",                   3, "show_text", "flag('battle_result') == 'win'", "击败灵虎后，你发现它守护的草丛中长满了珍稀灵草。", "", "", ""],
        ["exp_spirit_beast_4",   "",                   4, "give_resource", "flag('battle_result') == 'win'", "spirit_herb", "15", "", ""],
        ["exp_spirit_beast_5",   "",                   5, "show_text", "flag('battle_result') == 'lose'", "灵虎逼退了你，它慢慢躺回草丛中。你受了些伤，先退一步。", "", "", ""],
        # exp_boss_guardian：show_text → 战斗 → 灵石
        ["exp_boss_guardian_1",  "exp_boss_guardian", 1, "show_text", "", "碧鳞蟾蜍吐出一片毒雾，你感到一阵眩晕。此战避无可避！", "", "", ""],
        ["exp_boss_guardian_2",  "",                   2, "start_battle", "", "", "", "", ""],
        ["exp_boss_guardian_3",  "",                   3, "show_text", "flag('battle_result') == 'win'", "碧鳞蟾蜍倒地，一颗碧绿的内丹滚落在地，周围散落着大量灵石。", "", "", ""],
        ["exp_boss_guardian_4",  "",                   4, "give_resource", "flag('battle_result') == 'win'", "spirit_stone", "1000", "", ""],
        ["exp_boss_guardian_5",  "",                   5, "extraction", "flag('battle_result') == 'win'", "", "", "", ""],
        ["exp_boss_guardian_6",  "",                   6, "show_text", "flag('battle_result') == 'lose'", "你被击退，毒雾让身体虚弱。这次历练到此为止。", "", "", ""],
        ["exp_boss_guardian_7",  "",                   7, "extraction", "flag('battle_result') == 'lose'", "", "", "", ""],
        # === 扩充事件的动作 ===
        # exp_herb_field
        ["exp_herb_field_1", "exp_herb_field", 1, "show_text", "", "你小心采撷，收获颇丰。", "", "", ""],
        ["exp_herb_field_2", "",               2, "give_resource", "", "spirit_herb", "12", "", ""],
        # exp_old_hermit：选项（请教=悟性 / 论道=灵石）
        ["exp_old_hermit_1", "exp_old_hermit", 1, "show_text", "", "老者道：'年轻人，可愿与我手谈一局？'", "", "", ""],
        ["exp_old_hermit_2", "",               2, "show_options", "", "opt_hermit_learn", "opt_hermit_gift", "", ""],
        ["exp_old_hermit_3", "",               3, "show_text", "flag('choice') == 'opt_hermit_learn'", "老者指点了你几句修行心得，令你受益匪浅。", "", "", ""],
        ["exp_old_hermit_4", "",               4, "give_resource", "flag('choice') == 'opt_hermit_learn'", "spirit_stone", "200", "", ""],
        ["exp_old_hermit_5", "",               5, "show_text", "flag('choice') == 'opt_hermit_gift'", "老者赠你一袋灵草，飘然而去。", "", "", ""],
        ["exp_old_hermit_6", "",               6, "give_resource", "flag('choice') == 'opt_hermit_gift'", "spirit_herb", "20", "", ""],
        # exp_spirit_spring
        ["exp_spirit_spring_1", "exp_spirit_spring", 1, "show_text", "", "你掬一捧泉水饮下，灵气在经脉中流转，神清气爽。", "", "", ""],
        ["exp_spirit_spring_2", "",                  2, "give_resource", "", "spirit_stone", "250", "", ""],
        # exp_ruins_choice：选项（强闯=战斗+大奖 / 谨慎=小奖）
        ["exp_ruins_choice_1", "exp_ruins_choice", 1, "show_text", "", "洞府禁制隐隐发光，强闯有风险，但传承诱人。", "", "", ""],
        ["exp_ruins_choice_2", "",                  2, "show_options", "", "opt_ruins_force", "opt_ruins_careful", "", ""],
        ["exp_ruins_choice_3", "",                  3, "show_text", "flag('choice') == 'opt_ruins_force'", "你强行破禁，触发了守护傀儡！", "", "", ""],
        ["exp_ruins_choice_4", "",                  4, "start_battle", "flag('choice') == 'opt_ruins_force'", "foundation", "1", "", ""],
        ["exp_ruins_choice_5", "",                  5, "show_text", "flag('battle_result') == 'win'", "傀儡崩解，你取得了洞府中的宝物！", "", "", ""],
        ["exp_ruins_choice_6", "",                  6, "give_resource", "flag('battle_result') == 'win'", "spirit_stone", "800", "", ""],
        ["exp_ruins_choice_7", "",                  7, "give_resource", "flag('battle_result') == 'win'", "demon_core", "1", "", ""],
        ["exp_ruins_choice_8", "",                  8, "show_text", "flag('battle_result') == 'lose'", "傀儡太强，你只得退走。", "", "", ""],
        ["exp_ruins_choice_9", "",                  9, "show_text", "flag('choice') == 'opt_ruins_careful'", "你只取了外围零散之物，安全离开。", "", "", ""],
        ["exp_ruins_choice_10","",                  10, "give_resource", "flag('choice') == 'opt_ruins_careful'", "spirit_herb", "8", "", ""],
        # exp_wandering_cultivator：战斗
        ["exp_wandering_cultivator_1", "exp_wandering_cultivator", 1, "show_text", "", "散修抱拳：'得罪了！'", "", "", ""],
        ["exp_wandering_cultivator_2", "",                          2, "start_battle", "", "qi", "1", "", ""],
        ["exp_wandering_cultivator_3", "",                          3, "show_text", "flag('battle_result') == 'win'", "散修认输，留下灵石悻悻离去。", "", "", ""],
        ["exp_wandering_cultivator_4", "",                          4, "give_resource", "flag('battle_result') == 'win'", "spirit_stone", "400", "", ""],
        ["exp_wandering_cultivator_5", "",                          5, "show_text", "flag('battle_result') == 'lose'", "你技不如人，丢了些灵石。", "", "", ""],
        ["exp_wandering_cultivator_6", "",                          6, "give_resource", "flag('battle_result') == 'lose'", "spirit_stone", "-150", "", ""],
        # exp_poison_swamp
        ["exp_poison_swamp_1", "exp_poison_swamp", 1, "show_text", "", "你运转灵力抵御瘴气，勉强通过，消耗不小。", "", "", ""],
        ["exp_poison_swamp_2", "",                 2, "give_resource", "", "spirit_stone", "-100", "", ""],
        # exp_demon_ambush：群战
        ["exp_demon_ambush_1", "exp_demon_ambush", 1, "show_text", "", "三只金瞳妖豹气势汹汹！", "", "", ""],
        ["exp_demon_ambush_2", "",                 2, "start_battle", "", "golden", "3", "", ""],
        ["exp_demon_ambush_3", "",                 3, "show_text", "flag('battle_result') == 'win'", "妖豹尽数倒地，妖丹收入囊中！", "", "", ""],
        ["exp_demon_ambush_4", "",                 4, "give_resource", "flag('battle_result') == 'win'", "demon_core", "2", "", ""],
        ["exp_demon_ambush_5", "",                 5, "show_text", "flag('battle_result') == 'lose'", "寡不敌众，你负伤撤退。", "", "", ""],
        ["exp_demon_ambush_6", "",                 6, "extraction", "flag('battle_result') == 'lose'", "", "", "", ""],
        # exp_treasure_vault
        ["exp_treasure_vault_1", "exp_treasure_vault", 1, "show_text", "", "宝库中珍宝琳琅满目，你满载而归！", "", "", ""],
        ["exp_treasure_vault_2", "",                   2, "give_resource", "", "spirit_stone", "600", "", ""],
        ["exp_treasure_vault_3", "",                   3, "give_resource", "", "millennium_herb", "1", "", ""],
        # __sys_extraction__：撤离 → 返回
        ["sys_ext_1",            "__sys_extraction__", 1, "show_text", "", "带着此行收获，你踏上了归宗之路。", "", "", ""],
        ["sys_ext_2",            "",                   2, "extraction", "", "", "", "", ""],
    ]
    ws_append(wb.create_sheet("事件动作"), _ah, _ae, actions)

    # Sheet 3: 事件选项
    _oh = ["选项ID", "选项文本", "是否触发战斗", "敌方境界", "敌方数量", "奖励资源ID", "奖励数量", "备注"]
    _oe = ["option_id", "text_cn", "battle_trigger", "battle_enemy_realm", "battle_enemy_count", "reward_resource_id", "reward_amount", ""]
    options = [
        ["opt_cave_explore",   "勇闯洞穴", True,  "qi",         1, "", 0, ""],
        ["opt_cave_bypass",    "绕路前行", False, "",          0, "", 0, ""],
        ["opt_hermit_learn",   "请教修行", False, "",          0, "", 0, ""],
        ["opt_hermit_gift",    "讨要灵草", False, "",          0, "", 0, ""],
        ["opt_ruins_force",    "强闯洞府", True,  "foundation", 1, "", 0, ""],
        ["opt_ruins_careful",  "谨慎搜寻", False, "",          0, "", 0, ""],
    ]
    ws_append(wb.create_sheet("事件选项"), _oh, _oe, options)

    out = TABLE_DIR / "事件配置.xlsx"
    wb.save(out)
    print(f"[ok] {out.name}: {len(events)} events, {len(actions)} actions, {len(options)} options")


if __name__ == "__main__":
    gen()
    print("[done] 事件配表生成完毕")
